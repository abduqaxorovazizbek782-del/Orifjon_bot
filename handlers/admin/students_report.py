import io
from aiogram import Router, F
from aiogram.types import Message, BufferedInputFile
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from config import ADMINS
from database.engine import async_session
from database.models import Group
from keyboards.reply import admin_menu
from keyboards.buttons import BTN_STUDENTS_REPORT

router = Router()


@router.message(F.text == BTN_STUDENTS_REPORT, F.from_user.id.in_(ADMINS))
async def students_report(message: Message):
    async with async_session() as session:
        groups = (await session.scalars(
            select(Group).options(selectinload(Group.students))
            .order_by(Group.year, Group.name)
        )).all()

    if not groups:
        await message.answer("❌ Hozircha guruhlar yo'q.", reply_markup=admin_menu())
        return

    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    buffer = io.BytesIO()
    total_students = 0

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        any_sheet = False
        for g in groups:
            students = sorted(g.students, key=lambda s: (s.last_name or "", s.name or ""))
            if not students:
                continue
            any_sheet = True

            rows = []
            for i, s in enumerate(students, 1):
                rows.append({
                    "№": i,
                    "Ism": s.name,
                    "Familiya": s.last_name,
                    "Telefon": s.tel or "",
                    "Balans (so'm)": round(s.balance or 0),
                    "O'rtacha ball (%)": round(s.average_score or 0, 1),
                })
                total_students += 1

            df = pd.DataFrame(rows)
            sheet_name = f"{g.name} {g.year}"[:31]
            df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)

            ws = writer.sheets[sheet_name]

            ws.merge_cells("A1:F1")
            title = ws["A1"]
            title.value = (f"📗 {g.name} — {g.year}-o'quv yili  |  "
                           f"O'quvchilar: {len(students)}  |  "
                           f"Oylik: {g.monthly_price:,.0f} so'm")
            title.font = Font(bold=True, size=12, color="FFFFFF")
            title.fill = PatternFill("solid", fgColor="2E7D32")
            title.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 24

            header_fill = PatternFill("solid", fgColor="A5D6A7")
            thin = Side(style="thin", color="999999")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for col in range(1, 7):
                cell = ws.cell(row=3, column=col)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            for r in range(4, 4 + len(students)):
                for col in range(1, 7):
                    ws.cell(row=r, column=col).border = border

            widths = {"A": 5, "B": 18, "C": 18, "D": 16, "E": 16, "F": 18}
            for col, w in widths.items():
                ws.column_dimensions[col].width = w

        if not any_sheet:
            pd.DataFrame([{"Holat": "O'quvchilar yo'q"}]).to_excel(
                writer, index=False, sheet_name="Bo'sh")

    buffer.seek(0)
    file = BufferedInputFile(buffer.read(), filename="oquvchilar_hisoboti.xlsx")
    await message.answer_document(
        file,
        caption=(
            "📗 <b>O'quvchilar hisoboti</b>\n"
            "━━━━━━━━━━━━━━\n"
            f"🗂 Guruhlar: {len([g for g in groups if g.students])} ta\n"
            f"👥 Jami o'quvchilar: {total_students} ta\n\n"
            "Har bir guruh alohida varaqda, yil bo'yicha tartiblangan."
        ),
        reply_markup=admin_menu(),
    )
